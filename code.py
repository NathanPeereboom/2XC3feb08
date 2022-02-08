# code.py
import random
import timeit
from lab4 import *
from sorts import *
import math
from openpyxl import Workbook


# reused code from last assignment
def sortFunctionTimeTest(function, listt):
    start = timeit.default_timer()
    listt = function(listt)
    end = timeit.default_timer()
    return end - start


def randomAverageTimeTest(times,function,list_size):
    total = 0
    for i in range(times):
        listt = create_random_list(list_size)
        sortTime = sortFunctionTimeTest(function, listt)
        total += sortTime
    return total/int(times)


def bottomUpTest():
    # n=100
    ts = randomAverageTimeTest(500, mergesort, 100)
    # n=400
    tm = randomAverageTimeTest(500, mergesort, 400)
    # n=950
    tl = randomAverageTimeTest(500, mergesort, 950)

    # n=100
    bs = randomAverageTimeTest(500, mergesort_bottom, 100)
    # n=400
    bm = randomAverageTimeTest(500, mergesort_bottom, 400)
    # n=950
    bl = randomAverageTimeTest(500, mergesort_bottom, 950)
    
    wb = Workbook()
    sheet = wb.active
    sheet["A1"] = "List Size (n)"
    sheet["A2"] = "n = 100"
    sheet["A3"] = "n = 400"
    sheet["A4"] = "n = 950"

    sheet["B1"] = "Top-down mergesort"
    sheet["B2"] = ts
    sheet["B3"] = tm
    sheet["B4"] = tl

    sheet["C1"] = "Bottom-up mergesort"
    sheet["C2"] = bs
    sheet["C3"] = bm
    sheet["C4"] = bl

    wb.save(filename="topVdown data.xlsx")


def displayMergesort():
    mss = randomAverageTimeTest(500, mergesort, 100)
    # n=400
    msm = randomAverageTimeTest(500, mergesort, 400)
    # n=950
    msl = randomAverageTimeTest(500, mergesort, 950)

    m3s = randomAverageTimeTest(500, mergesort_three, 100)
    # n=400
    m3m = randomAverageTimeTest(500, mergesort_three, 400)
    # n=950
    m3l = randomAverageTimeTest(500, mergesort_three, 950)

    wb = Workbook()
    sheet = wb.active
    sheet["A1"] = "List Size (n)"
    sheet["A2"] = "n = 100"
    sheet["A3"] = "n = 400"
    sheet["A4"] = "n = 950"

    sheet["B1"] = "Mergesort"
    sheet["B2"] = mss
    sheet["B3"] = msm
    sheet["B4"] = msl

    sheet["C1"] = "Three-way mergesort"
    sheet["C2"] = m3s
    sheet["C3"] = m3m
    sheet["C4"] = m3l

    wb.save(filename="mergesort data.xlsx")

# mergesort vs factor
# factor between 0 and 0.5

# column A is average case performance, coulmn B is worst case performance and column C is n
def worstCase():
    bestMergesort = mergesort  # replace if another mergesort runs better than the standard

    num = 1
    workbook = Workbook()
    sheet = workbook.active
    factor = 0

    # headers
    sheet["A1"] = "Factor"
    sheet["B1"] = "Mergesort Runtime"

    sheet["C1"] = "Mergesort 3-Way Runtime"

    factor = 0
    times = 2
    listLength = 100
    while factor <= 0.5:
        spot = str("B" + str(times))
        runTime = sortFunctionTimeTest(bestMergesort, create_near_sorted_list(listLength, factor))
        sheet[spot] = runTime

        spot = str("C" + str(times))
        runTime = sortFunctionTimeTest(mergesort_three, create_near_sorted_list(listLength, factor))
        sheet[spot] = runTime

        spot = str("A" + str(times))
        sheet[spot] = factor
        factor += 0.1
        times += 1

    workbook.save(filename="Lab 4 Worst Case.xlsx")


# Test Test
displayMergesort()
worstCase()
bottomUpTest()
